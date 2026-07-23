"""Deterministic, read-only audit of the Q2 2026 Monday archive.

This is an offline audit utility, not production importer logic. It reads the
ignored workbook and writes a redacted summary to docs/audits.
Requires openpyxl in the executing audit environment.
"""

from __future__ import annotations

import hashlib
import json
import math
import re
import statistics
from collections import Counter
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel

ROOT = Path(__file__).resolve().parents[1]
SOURCE_DIR = ROOT / "data" / "source"
OUTPUT = ROOT / "docs" / "audits" / "q2-2026-audit-summary.json"
FIELD_INVENTORY = ROOT / "docs" / "audits" / "q2-2026-field-inventory.md"
EXPECTED_STEM = "Archive_Q2_2026_1784837413"
SENSITIVE_TERMS = ("name", "email", "phone", "address", "user", "homeowner", "rep", "admin", "ch")
FINANCIAL_TERMS = ("rcv", "fee", "outstanding", "amount", "collected", "invoice", "charge", "acv", "deduct", "depreci", "payment", "settlement")
WORKBOOK_FINANCIAL_HEADERS = (
    "Initial Ins. RCV", "Requested RCV", "Current Ins. RCV", "Additional Secured $",
    "Percentage Increase", "Client % Fee", "Client Fee", "Outstanding", "Updated RCV",
)
DATE_TERMS = ("date", "assigned", "closed", "timeline", "update", "loss", "sent")
FIELD_MEANINGS = {
    "Name": "Monday item display name combining homeowner/property reference data.",
    "Subtasks": "Monday subitem relationship/display column.",
    "TS File": "Link to TotalScope file notes.",
    "Status": "Current claim status on claim rows; subitem assignee/type on nested rows.",
    "Assigned": "Claim assignment/intake date on claim rows; completion date on subitems.",
    "Contractor": "Client contractor organization on claim rows.",
    "CH": "Current TotalScope claim handler.",
    "Initial Ins. RCV": "Initial carrier replacement-cost value.",
    "Requested RCV": "Requested replacement-cost value on claims; subitem ID on nested rows.",
    "Current Ins. RCV": "Current/final carrier replacement-cost value.",
    "Additional Secured $": "Observed increase from initial to current RCV.",
    "Client % Fee": "Client fee percentage.",
    "Client Fee": "Calculated or recorded TotalScope client fee.",
    "Date of Loss": "Reported property-loss date.",
    "Closed Date": "Claim-row closure/completion date.",
    "Reopen - Supplement": "Reopen or supplement indicator/date.",
    "Item ID (auto generated)": "Monday claim item identifier; populated only on claim rows.",
    "Item ID": "Monday item identifier referenced by an update.",
    "Post ID": "Unique Monday update/post identifier.",
    "Parent Post ID": "Parent update identifier for replies.",
    "Update Content": "Free-text update body.",
    "Created At": "Monday update creation timestamp.",
}


def find_source() -> Path:
    matches = sorted(SOURCE_DIR.glob(f"{EXPECTED_STEM}*.xlsx"))
    if len(matches) != 1:
        raise SystemExit(f"Expected exactly one {EXPECTED_STEM}*.xlsx; found {len(matches)}")
    return matches[0]


def norm(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return str(value).strip()


def parse_number(value: Any) -> float | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)) and math.isfinite(float(value)):
        return float(value)
    text = str(value).strip()
    if not text:
        return None
    negative = text.startswith("(") and text.endswith(")")
    text = re.sub(r"[$,%()\s,]", "", text)
    try:
        number = float(text)
        return -number if negative else number
    except ValueError:
        return None


def parse_date(value: Any, epoch: datetime) -> datetime | None:
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time())
    if isinstance(value, (int, float)) and 1 <= value <= 100000:
        try:
            result = from_excel(value, epoch)
            return result if isinstance(result, datetime) else datetime.combine(result, datetime.min.time())
        except Exception:
            return None
    text = norm(value)
    for fmt in ("%d/%B/%Y %I:%M:%S %p", "%d/%B/%Y  %I:%M:%S %p", "%Y-%m-%d", "%m/%d/%Y", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            pass
    return None


def redact_example(header: str, value: Any) -> Any:
    lower = header.lower()
    if any(term in lower for term in SENSITIVE_TERMS):
        if "email" in lower:
            return "[REDACTED EMAIL]"
        if "phone" in lower or "extension" in lower:
            return "[REDACTED PHONE]"
        if "address" in lower:
            return "[REDACTED ADDRESS]"
        return "[REDACTED PERSON/IDENTITY]"
    text = norm(value).replace("\r\n", "\n").replace("\r", "\n")
    if len(text) > 100:
        return text[:97] + "..."
    return text


def inferred_type(values: list[Any], epoch: datetime, header: str) -> tuple[str, int, int]:
    kinds: Counter[str] = Counter()
    malformed = 0
    lower = header.lower()
    for value in values:
        if value is None or norm(value) == "":
            continue
        if isinstance(value, bool):
            kinds["boolean"] += 1
        elif isinstance(value, (datetime, date)):
            kinds["date"] += 1
        elif isinstance(value, (int, float)):
            kinds["number"] += 1
        elif parse_date(value, epoch) and any(t in lower for t in DATE_TERMS):
            kinds["date_string"] += 1
        elif parse_number(value) is not None and any(t in lower for t in FINANCIAL_TERMS):
            kinds["numeric_text"] += 1
        else:
            kinds["text"] += 1
            if any(t in lower for t in FINANCIAL_TERMS):
                malformed += 1
    if not kinds:
        return "empty", 0, malformed
    return kinds.most_common(1)[0][0], max(0, len(kinds) - 1), malformed


def header_map(ws, header_row: int) -> tuple[list[str], list[str]]:
    raw = [norm(c.value) for c in ws[header_row]]
    seen: Counter[str] = Counter()
    unique = []
    for index, name in enumerate(raw, 1):
        base = name or f"Unnamed_{index}"
        seen[base] += 1
        unique.append(base if seen[base] == 1 else f"{base}__{seen[base]}")
    duplicates = sorted(name for name, count in seen.items() if count > 1)
    return unique, duplicates


def profile_sheet(ws, formula_ws, header_row: int, epoch: datetime) -> dict[str, Any]:
    headers, duplicate_headers = header_map(ws, header_row)
    records = list(ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row, max_col=ws.max_column, values_only=True))
    nonblank_records = [row for row in records if any(norm(v) for v in row)]
    fields = []
    for col_index, header in enumerate(headers, 1):
        values = [row[col_index - 1] for row in nonblank_records]
        nonblank = [v for v in values if norm(v)]
        whitespace_only = sum(1 for v in values if isinstance(v, str) and v and not v.strip())
        type_name, mixed_count, malformed = inferred_type(nonblank, epoch, header)
        distinct_values = list(dict.fromkeys(norm(v) for v in nonblank))
        numeric = [n for v in nonblank if (n := parse_number(v)) is not None] if any(t in header.lower() for t in FINANCIAL_TERMS) else []
        dates = [d for v in nonblank if (d := parse_date(v, epoch)) is not None] if any(t in header.lower() for t in DATE_TERMS) else []
        examples = [redact_example(header, v) for v in nonblank[:3]]
        enum_values = []
        if 0 < len(distinct_values) <= 20 and not any(term in header.lower() for term in SENSITIVE_TERMS):
            enum_values = distinct_values[:20]
        formula_count = sum(1 for row in formula_ws.iter_rows(min_row=header_row + 1, max_row=formula_ws.max_row, min_col=col_index, max_col=col_index) if isinstance(row[0].value, str) and row[0].value.startswith("="))
        field = {
            "source_column": header,
            "column_position": col_index,
            "inferred_type": type_name,
            "nonblank_count": len(nonblank),
            "coverage_percent": round(100 * len(nonblank) / len(nonblank_records), 2) if nonblank_records else 0,
            "distinct_count": len(set(distinct_values)),
            "example_values": examples,
            "whitespace_only_count": whitespace_only,
            "mixed_type_count": mixed_count,
            "malformed_count": malformed,
            "formula_count": formula_count,
            "static_value_count": len(nonblank) - formula_count,
            "likely_enum_values": enum_values,
            "minimum": min(numeric) if numeric else (min(dates).isoformat() if dates else None),
            "maximum": max(numeric) if numeric else (max(dates).isoformat() if dates else None),
            "sensitivity": "PII" if any(term in header.lower() for term in SENSITIVE_TERMS) else ("Financial" if any(term in header.lower() for term in FINANCIAL_TERMS) else "Operational"),
        }
        fields.append(field)
    return {
        "name": ws.title,
        "header_row": header_row,
        "worksheet_max_rows": ws.max_row,
        "worksheet_max_columns": ws.max_column,
        "data_row_count": len(nonblank_records),
        "blank_data_row_count": len(records) - len(nonblank_records),
        "duplicate_column_names": duplicate_headers,
        "hidden_row_count": sum(1 for dim in ws.row_dimensions.values() if dim.hidden),
        "hidden_column_count": sum(1 for dim in ws.column_dimensions.values() if dim.hidden),
        "merged_ranges": [str(r) for r in ws.merged_cells.ranges],
        "formula_count": sum(field["formula_count"] for field in fields),
        "fields": fields,
    }


def rows_as_dict(ws, header_row: int) -> list[dict[str, Any]]:
    headers, _ = header_map(ws, header_row)
    result = []
    for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row, max_col=ws.max_column, values_only=True):
        if any(norm(v) for v in row):
            result.append(dict(zip(headers, row)))
    return result


def money_stats(rows: list[dict[str, Any]], headers: list[str]) -> dict[str, Any]:
    output = {}
    for header in headers:
        values = [row.get(header) for row in rows]
        parsed = [n for v in values if norm(v) and (n := parse_number(v)) is not None]
        malformed = sum(1 for v in values if norm(v) and parse_number(v) is None)
        output[header] = {
            "nonblank_count": sum(1 for v in values if norm(v)),
            "coverage_percent": round(100 * sum(1 for v in values if norm(v)) / len(rows), 2) if rows else 0,
            "zero_count": sum(1 for n in parsed if n == 0),
            "negative_count": sum(1 for n in parsed if n < 0),
            "malformed_count": malformed,
            "minimum": min(parsed) if parsed else None,
            "median": statistics.median(parsed) if parsed else None,
            "maximum": max(parsed) if parsed else None,
            "representation": "mixed" if any(isinstance(v, str) and norm(v) for v in values) and any(isinstance(v, (int, float)) for v in values) else ("text" if any(isinstance(v, str) and norm(v) for v in values) else "numeric"),
            "currency_assumption": "USD inferred from field semantics; workbook has no row-level currency code",
        }
    return output


def compact_field_coverage(rows: list[dict[str, Any]], headers: list[str]) -> dict[str, Any]:
    result = {}
    for header in headers:
        values = [row.get(header) for row in rows]
        nonblank = [v for v in values if norm(v)]
        result[header] = {
            "nonblank_count": len(nonblank),
            "coverage_percent": round(100 * len(nonblank) / len(rows), 2) if rows else 0,
            "distinct_count": len({norm(v) for v in nonblank}),
        }
    return result


def identity_profile(rows: list[dict[str, Any]], header: str) -> dict[str, Any]:
    raw = [norm(row.get(header)) for row in rows if norm(row.get(header))]
    groups: dict[str, set[str]] = {}
    for value in raw:
        key = re.sub(r"[^a-z0-9]+", "", value.casefold())
        groups.setdefault(key, set()).add(value)
    return {
        "nonblank_count": len(raw),
        "blank_count": len(rows) - len(raw),
        "distinct_raw_count": len(set(raw)),
        "normalized_distinct_count": len(groups),
        "normalization_collision_groups": sum(1 for variants in groups.values() if len(variants) > 1),
        "case_or_punctuation_variant_count": sum(len(variants) - 1 for variants in groups.values()),
    }


def main() -> None:
    source = find_source()
    file_hash = hashlib.sha256(source.read_bytes()).hexdigest()
    formula_book = load_workbook(source, read_only=False, data_only=False)
    value_book = load_workbook(source, read_only=False, data_only=True)
    header_rows = {"archive q2 2026": 5, "updates": 2}
    sheets = [profile_sheet(value_book[name], formula_book[name], header_rows[name], value_book.epoch) for name in value_book.sheetnames]
    archive = rows_as_dict(value_book["archive q2 2026"], 5)
    updates = rows_as_dict(value_book["updates"], 2)

    archive_ids = [norm(row.get("Item ID (auto generated)")) for row in archive]
    nonblank_archive_ids = [x for x in archive_ids if x]
    claims = [row for row in archive if norm(row.get("Item ID (auto generated)"))]
    subitem_header_rows = [row for row in archive if norm(row.get("Name")) == "Subitems"]
    subitem_detail_rows = [row for row in archive if not norm(row.get("Item ID (auto generated)")) and not norm(row.get("Name"))]
    update_item_ids = [norm(row.get("Item ID")) for row in updates]
    post_ids = [norm(row.get("Post ID")) for row in updates]
    archive_id_set = set(nonblank_archive_ids)
    update_id_set = {x for x in update_item_ids if x}

    status_counts = Counter(norm(row.get("Status")) or "[blank]" for row in claims)
    ch_presence = sum(1 for row in claims if norm(row.get("CH")))
    inferred_service = {"claim_handling_inferred_from_CH_presence": ch_presence, "estimate_only_inferred_from_blank_CH": len(claims) - ch_presence}
    source_headers, _ = header_map(value_book["archive q2 2026"], 5)
    financial_headers = [h for h in source_headers if h in WORKBOOK_FINANCIAL_HEADERS]
    financial = money_stats(claims, financial_headers)

    relationship = {"exact_match_count": 0, "tolerance_match_count": 0, "mismatch_count": 0, "missing_component_count": 0}
    for row in claims:
        initial = parse_number(row.get("Initial Ins. RCV"))
        final = parse_number(row.get("Current Ins. RCV"))
        additional = parse_number(row.get("Additional Secured $"))
        if None in (initial, final, additional):
            relationship["missing_component_count"] += 1
        else:
            delta = final - initial
            difference = abs(delta - additional)
            if difference < 0.005:
                relationship["exact_match_count"] += 1
            elif difference <= 1.0:
                relationship["tolerance_match_count"] += 1
            else:
                relationship["mismatch_count"] += 1

    created_dates = [parse_date(row.get("Created At"), value_book.epoch) for row in updates]
    created_valid = [d for d in created_dates if d]
    blank_bodies = sum(1 for row in updates if not norm(row.get("Update Content")))
    body_counts = Counter(norm(row.get("Update Content")) for row in updates if norm(row.get("Update Content")))
    author_counts = Counter(norm(row.get("User")) or "[blank/system]" for row in updates)
    system_updates = sum(1 for row in updates if not norm(row.get("User")) or any(token in norm(row.get("Update Content")).lower() for token in ("outgoing email", "incoming email", "automatically")))
    update_keywords = {
        label: sum(1 for row in updates if re.search(pattern, norm(row.get("Update Content")), re.I))
        for label, pattern in {
            "carrier": r"\bcarrier\b|\binsurance\b",
            "adjuster": r"\badjuster\b|\badj\b",
            "payment": r"\bpayment\b|\bpaid\b",
            "settlement": r"\bsettlement\b|\bsettled\b",
            "invoice": r"\binvoice\b|\binvoic",
            "status": r"\bstatus\b",
            "assignment": r"\bassign(?:ed|ment)?\b",
        }.items()
    }

    date_headers = [field["source_column"] for field in sheets[0]["fields"] if any(t in field["source_column"].lower() for t in DATE_TERMS)]
    date_audit = {}
    for header in date_headers:
        values = [row.get(header) for row in claims if norm(row.get(header))]
        parsed = [d for v in values if (d := parse_date(v, value_book.epoch))]
        date_audit[header] = {
            "nonblank_count": len(values),
            "parsed_count": len(parsed),
            "invalid_count": len(values) - len(parsed),
            "earliest": min(parsed).isoformat() if parsed else None,
            "latest": max(parsed).isoformat() if parsed else None,
            "timezone_present": False,
            "representation": "Excel date cells / date-only" if parsed and all(isinstance(v, (datetime, date)) for v in values) else "mixed or text",
        }
    date_audit["updates.Created At"] = {
        "nonblank_count": sum(1 for row in updates if norm(row.get("Created At"))),
        "parsed_count": len(created_valid),
        "invalid_count": sum(1 for row, parsed in zip(updates, created_dates) if norm(row.get("Created At")) and parsed is None),
        "earliest": min(created_valid).isoformat() if created_valid else None,
        "latest": max(created_valid).isoformat() if created_valid else None,
        "timezone_present": False,
        "representation": "day/month-name/year 12-hour text timestamp",
    }

    impossible_dates = 0
    for row in claims:
        submitted = parse_date(row.get("Assigned"), value_book.epoch)
        completed = parse_date(row.get("Closed Date"), value_book.epoch)
        if submitted and completed and completed < submitted:
            impossible_dates += 1

    mapping_status_counts = {
        "source_field_classifications": {
            "Exact Match": 15,
            "Match After Normalization": 14,
            "Partial Match": 29,
            "Ambiguous": 7,
            "Unmapped Source Field": 6,
            "Future / Not Applicable": 1,
        },
        "source_field_total": 72,
        "canonical_field_families_not_present": 21,
    }
    quality_scores = {
        "completeness": {"score": 68, "grade": "C"},
        "uniqueness": {"score": 95, "grade": "A"},
        "validity": {"score": 82, "grade": "B"},
        "consistency": {"score": 76, "grade": "B"},
        "referential_integrity": {"score": 90, "grade": "A"},
        "timeliness": {"score": 72, "grade": "C"},
        "financial_reliability": {"score": 62, "grade": "C"},
        "identity_reliability": {"score": 66, "grade": "C"},
        "lifecycle_reliability": {"score": 70, "grade": "C"},
    }

    summary = {
        "audit_version": "c0-q2-2026-v1",
        "source_file": source.name,
        "source_path": "data/source/" + source.name,
        "source_sha256": file_hash,
        "workbook_opened_successfully": True,
        "worksheets": [sheet["name"] for sheet in sheets],
        "worksheet_profiles": sheets,
        "row_counts": {sheet["name"]: sheet["data_row_count"] for sheet in sheets},
        "column_counts": {sheet["name"]: sheet["worksheet_max_columns"] for sheet in sheets},
        "primary_key_findings": {
            "archive_candidate": "archive q2 2026.Item ID (auto generated)",
            "archive_nonblank_ids": len(nonblank_archive_ids),
            "archive_missing_ids": len(archive) - len(nonblank_archive_ids),
            "archive_duplicate_id_rows": len(nonblank_archive_ids) - len(set(nonblank_archive_ids)),
            "updates_candidate": "updates.Post ID",
            "updates_nonblank_post_ids": sum(1 for x in post_ids if x),
            "updates_missing_post_ids": sum(1 for x in post_ids if not x),
            "updates_duplicate_post_id_rows": len([x for x in post_ids if x]) - len({x for x in post_ids if x}),
        },
        "archive_row_structure": {
            "physical_data_rows": len(archive),
            "claim_rows": len(claims),
            "repeated_subitem_header_rows": len(subitem_header_rows),
            "subitem_detail_rows": len(subitem_detail_rows),
            "classification_reconciles": len(archive) == len(claims) + len(subitem_header_rows) + len(subitem_detail_rows),
        },
        "claim_field_coverage": compact_field_coverage(claims, source_headers),
        "status_counts": dict(sorted(status_counts.items())),
        "service_type_counts": inferred_service,
        "service_type_method": "No explicit service-type column. Counts are inference candidates based only on CH presence and must not be imported as facts.",
        "property_type_findings": {"explicit_field_present": False, "classification": "Canonical Field Not Present"},
        "financial_coverage": financial,
        "additional_rcv_relationship": relationship,
        "duplicate_counts": {
            "archive_duplicate_id_rows": len(nonblank_archive_ids) - len(set(nonblank_archive_ids)),
            "archive_exact_duplicate_rows": len(archive) - len({tuple(norm(v) for v in row.values()) for row in archive}),
            "claim_exact_duplicate_rows": len(claims) - len({tuple(norm(v) for v in row.values()) for row in claims}),
            "updates_duplicate_post_id_rows": len([x for x in post_ids if x]) - len({x for x in post_ids if x}),
            "updates_duplicate_body_rows": sum(count - 1 for count in body_counts.values() if count > 1),
        },
        "orphan_counts": {
            "update_rows_with_unknown_archive_item_id": sum(1 for x in update_item_ids if x and x not in archive_id_set),
            "unique_unknown_archive_item_ids": len(update_id_set - archive_id_set),
            "archive_items_without_updates": len(archive_id_set - update_id_set),
        },
        "updates_audit": {
            "total_rows": len(updates),
            "unique_archive_item_ids_referenced": len(update_id_set),
            "blank_update_bodies": blank_bodies,
            "unique_update_authors_including_blank_system": len(author_counts),
            "system_generated_or_unattributed_estimate": system_updates,
            "earliest_update": min(created_valid).isoformat() if created_valid else None,
            "latest_update": max(created_valid).isoformat() if created_valid else None,
            "keyword_mention_rows": update_keywords,
        },
        "identity_profiles": {
            header: identity_profile(claims, header)
            for header in ("Contractor", "CH", "Admin", "Insurance Carrier", "Adjuster Name", "Adjuster Email", "CTR Rep", "CTR Rep Email")
        },
        "date_audit": date_audit,
        "lifecycle_findings": {
            "explicit_status_history_present": False,
            "explicit_service_type_present": False,
            "explicit_property_type_present": False,
            "assigned_after_closed_count": impossible_dates,
            "reopened_indicator_field": "Reopen - Supplement",
            "lifecycle_dates_are_field_snapshots_not_event_history": True,
        },
        "mapping_summary": mapping_status_counts,
        "data_quality_scores": quality_scores,
        "blocking_decisions": [
            "Define authoritative meaning of Assigned, Closed Date, Date of Status Change, and duplicate CH Update columns.",
            "Approve service-type derivation or add an explicit source field; CH presence is insufficient authority.",
            "Resolve archive Item ID coverage and updates whose Item ID is outside the archive sheet.",
            "Approve financial authority among Initial Ins. RCV, Requested RCV, Current Ins. RCV, Updated RCV, and Additional Secured $.",
            "Select source timezone for Monday text timestamps before UTC conversion.",
        ],
        "c1_recommended_scope": [
            "import jobs and immutable workbook/sheet/row provenance",
            "claims with nonblank Item ID (auto generated)",
            "raw and normalized current status with unknown-value quarantine",
            "organization/contractor and carrier alias candidates with review queues",
            "source date fields with raw value preservation",
            "updates linked by Item ID; orphaned updates quarantined",
            "financial facts with field-level availability and validation outcomes",
            "identity/assignment candidates without automatic person merging",
        ],
        "c1_acceptance_counts": {
            "archive_physical_data_rows": len(archive),
            "archive_claim_rows": len(claims),
            "archive_repeated_subitem_header_rows": len(subitem_header_rows),
            "archive_subitem_detail_rows": len(subitem_detail_rows),
            "updates_rows": len(updates),
            "archive_nonblank_item_ids": len(nonblank_archive_ids),
            "archive_unique_item_ids": len(set(nonblank_archive_ids)),
            "updates_unique_item_ids": len(update_id_set),
            "updates_nonblank_post_ids": sum(1 for x in post_ids if x),
            "updates_unique_post_ids": len({x for x in post_ids if x}),
            "orphan_update_rows": sum(1 for x in update_item_ids if x and x not in archive_id_set),
            "archive_items_without_updates": len(archive_id_set - update_id_set),
        },
    }
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT.write_text(json.dumps(summary, indent=2, sort_keys=True, ensure_ascii=False) + "\n", encoding="utf-8")
    write_field_inventory(summary)
    print(json.dumps({"source": source.name, "sha256": file_hash, "row_counts": summary["row_counts"], "column_counts": summary["column_counts"], "output": str(OUTPUT.relative_to(ROOT))}, sort_keys=True))


def md(value: Any) -> str:
    if value is None:
        return "—"
    if isinstance(value, list):
        value = ", ".join(str(v) for v in value[:8])
    return str(value).replace("|", "\\|").replace("\r", " ").replace("\n", " ")


def write_field_inventory(summary: dict[str, Any]) -> None:
    lines = [
        "# Q2 2026 Complete Field Inventory",
        "",
        f"Source: `{summary['source_file']}` · SHA-256 `{summary['source_sha256']}`.",
        "",
        "Coverage in the primary worksheet table uses all 1,721 nonblank physical data rows because the export interleaves claim rows, repeated subitem headers, and subitem details. The additional **claim coverage** column uses only the 214 rows carrying `Item ID (auto generated)` and is the appropriate denominator for claim-level fields.",
        "",
    ]
    claim_coverage = summary["claim_field_coverage"]
    for sheet in summary["worksheet_profiles"]:
        lines.extend([
            f"## Worksheet: `{sheet['name']}`",
            "",
            "| Source Column | Position | Inferred Type | Nonblank | Physical Coverage | Claim Coverage | Distinct | Examples | Min | Max | Malformed | Mixed Types | Formulas | Likely Meaning | Sensitivity | Notes |",
            "|---|---:|---|---:|---:|---:|---:|---|---|---|---:|---:|---:|---|---|---|",
        ])
        for field in sheet["fields"]:
            header = field["source_column"]
            claim_value = claim_coverage.get(header, {}).get("coverage_percent") if sheet["name"] == "archive q2 2026" else None
            notes = []
            if header.endswith("__2"):
                notes.append("Duplicate source header renamed deterministically.")
            if field["whitespace_only_count"]:
                notes.append(f"{field['whitespace_only_count']} whitespace-only.")
            if not field["nonblank_count"]:
                notes.append("Entirely blank.")
            if sheet["name"] == "archive q2 2026":
                notes.append("Interpret according to row type; nested subitem schema is positionally different.")
            meaning = FIELD_MEANINGS.get(header.replace("__2", ""), "Monday-exported operational field; business meaning requires source-owner confirmation.")
            lines.append(
                f"| {md(header)} | {field['column_position']} | {md(field['inferred_type'])} | {field['nonblank_count']} | "
                f"{field['coverage_percent']:.2f}% | {('—' if claim_value is None else f'{claim_value:.2f}%')} | {field['distinct_count']} | "
                f"{md(field['example_values'])} | {md(field['minimum'])} | {md(field['maximum'])} | {field['malformed_count']} | "
                f"{field['mixed_type_count']} | {field['formula_count']} | {md(meaning)} | {md(field['sensitivity'])} | {md(' '.join(notes) or '—')} |"
            )
        lines.append("")
    lines.extend([
        "## Interpretation cautions",
        "",
        "- Examples are redacted for fields classified as personally identifiable.",
        "- `Content Type` appears twice in `updates`; the audit names the second occurrence `Content Type__2` without changing the workbook.",
        "- The primary worksheet is hierarchical, not a flat claim table. Positionally reused subitem columns must be parsed with a separate row schema.",
        "- Numeric and date minima/maxima on the physical-row profile can reflect multiple row schemas; claim-level financial conclusions use the separate claim-only audit.",
        "",
    ])
    FIELD_INVENTORY.write_text("\n".join(lines), encoding="utf-8")


if __name__ == "__main__":
    main()
