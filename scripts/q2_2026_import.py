#!/usr/bin/env python3
"""Deterministic, read-only parser and Supabase importer for the Q2 2026 Monday archive."""
from __future__ import annotations

import argparse
import hashlib
import json
import os
import re
import sys
import urllib.error
import urllib.request
import urllib.parse
from collections import Counter
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
SOURCE = ROOT / "data/source/Archive_Q2_2026_1784837413 (1).xlsx"
EXPECTED_SHA256 = "8c25be883993f821e6deff6a6aa787d30ee9d794b5cf7fe73a41b58c67f06323"
IMPORTER_VERSION = "c1-q2-2026-v1"
TRANSFORMATION_VERSION = "monday-archive-q2-2026-v1"
DATABASE_BATCH_SIZES = {
    "source_rows": 250,
    "claims": 25,
    "subitem_headers": 100,
    "subitem_details": 100,
    "updates": 250,
}
ARCHIVE_SHEET = "archive q2 2026"
UPDATES_SHEET = "updates"
EXPECTED = {
    "claims": 214, "complete": 177, "closed": 37, "subitem_headers": 148,
    "subitem_details": 1359, "updates": 5957, "unmatched_update_rows": 58,
    "unmatched_update_item_ids": 56, "unique_post_ids": 5957,
    "blank_update_bodies": 12, "duplicate_body_surplus": 398,
    "exact_matches": 213, "tolerance_matches": 0, "mismatches": 0,
    "missing_components": 1,
}


def text(value: Any) -> str | None:
    if value is None:
        return None
    result = str(value).strip()
    return result or None


def json_value(value: Any) -> Any:
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, Decimal):
        return str(value)
    return value


def decimal_value(value: Any) -> str | None:
    if value is None or (isinstance(value, str) and not value.strip()):
        return None
    try:
        cleaned = re.sub(r"[$,%\s,]", "", str(value))
        return str(Decimal(cleaned))
    except (InvalidOperation, ValueError):
        return None


def disambiguated_headers(sheet, header_row: int) -> list[dict[str, Any]]:
    counts: Counter[str] = Counter()
    result = []
    for position, cell in enumerate(sheet[header_row], 1):
        raw = text(cell.value) or f"unnamed_column_{position}"
        counts[raw] += 1
        internal = raw if counts[raw] == 1 else f"{raw}__column_{position}"
        result.append({"position": position, "raw": raw, "internal": internal})
    return result


def row_object(values: tuple[Any, ...], headers: list[dict[str, Any]]) -> dict[str, Any]:
    return {header["internal"]: json_value(values[index]) for index, header in enumerate(headers)}


def raw_date(value: Any) -> dict[str, Any]:
    raw = json_value(value)
    if raw is None or raw == "":
        return {"raw": raw, "parsed": None, "timezone_status": "date_only"}
    if isinstance(value, datetime):
        return {"raw": raw, "parsed": value.isoformat(), "timezone_status": "date_only" if value.time().isoformat() == "00:00:00" else "unknown_timezone"}
    if isinstance(value, date):
        return {"raw": raw, "parsed": value.isoformat(), "timezone_status": "date_only"}
    return {"raw": raw, "parsed": None, "timezone_status": "unparseable"}


def parse_update_timestamp(value: Any) -> dict[str, Any]:
    raw = text(value)
    if not raw:
        return {"raw": raw, "parsed": None, "timezone_status": "unknown_timezone"}
    for fmt in ("%d %B %Y %I:%M:%S %p", "%d %B %Y %I:%M %p"):
        try:
            parsed = datetime.strptime(raw, fmt)
            return {"raw": raw, "parsed": parsed.isoformat(), "timezone_status": "unknown_timezone"}
        except ValueError:
            continue
    return {"raw": raw, "parsed": None, "timezone_status": "unparseable"}


def normalized_alias(value: Any) -> str | None:
    raw = text(value)
    return re.sub(r"[^a-z0-9]+", " ", raw.lower()).strip() if raw else None


def fingerprint_body(value: Any) -> str | None:
    body = text(value)
    if not body:
        return None
    return hashlib.sha256(body.encode("utf-8")).hexdigest()


def parse_workbook(path: Path = SOURCE, allow_checksum_override: bool = False) -> dict[str, Any]:
    if not path.is_file():
        raise FileNotFoundError(f"Workbook not found: {path}")
    source_bytes = path.read_bytes()
    source_hash = hashlib.sha256(source_bytes).hexdigest()
    if source_hash != EXPECTED_SHA256 and not allow_checksum_override:
        raise ValueError(f"Workbook SHA-256 mismatch: expected {EXPECTED_SHA256}, received {source_hash}")

    workbook = load_workbook(path, read_only=False, data_only=True)
    if workbook.sheetnames != [ARCHIVE_SHEET, UPDATES_SHEET]:
        raise ValueError(f"Unexpected worksheets: {workbook.sheetnames}")
    archive = workbook[ARCHIVE_SHEET]
    updates_sheet = workbook[UPDATES_SHEET]
    if (archive.max_row, archive.max_column) != (1726, 61):
        raise ValueError(f"Unexpected archive dimensions: {archive.max_row}x{archive.max_column}")
    if (updates_sheet.max_row, updates_sheet.max_column) != (5959, 11):
        raise ValueError(f"Unexpected updates dimensions: {updates_sheet.max_row}x{updates_sheet.max_column}")

    archive_headers = disambiguated_headers(archive, 5)
    update_headers = disambiguated_headers(updates_sheet, 2)
    source_rows, claims, subitem_headers, subitem_details = [], [], [], []
    preceding_claim_id = None
    preceding_header_row = None

    for physical_row, values in enumerate(
        archive.iter_rows(min_row=6, max_row=1726, max_col=61, values_only=True), 6
    ):
        row = row_object(values, archive_headers)
        item_id = text(row.get("Item ID (auto generated)"))
        name = text(row.get("Name"))
        if item_id:
            row_type = "claim"
            preceding_claim_id = item_id
            preceding_header_row = None
        elif name == "Subitems":
            row_type = "subitem_header"
            preceding_header_row = physical_row
        else:
            row_type = "subitem_detail"
        source_row = {
            "worksheet": ARCHIVE_SHEET, "physical_row": physical_row,
            "row_type": row_type, "source_item_id": item_id,
            "raw": row, "normalized": row,
        }
        source_rows.append(source_row)
        if row_type == "claim":
            facts = {}
            for position, source, metric in (
                (9, "Initial Ins. RCV", "initial_rcv"),
                (10, "Requested RCV", "requested_rcv"),
                (11, "Current Ins. RCV", "current_rcv"),
                (12, "Additional Secured $", "additional_secured"),
                (15, "Client Fee", "client_fee"),
                (13, "Percentage Increase", "percentage_increase"),
            ):
                raw = row.get(source)
                parsed = decimal_value(raw)
                facts[metric] = {
                    "source_field": source, "source_column": position, "raw": json_value(raw),
                    "parsed": parsed, "availability": "captured" if parsed is not None else "not_captured",
                }
            initial = Decimal(facts["initial_rcv"]["parsed"]) if facts["initial_rcv"]["parsed"] is not None else None
            current = Decimal(facts["current_rcv"]["parsed"]) if facts["current_rcv"]["parsed"] is not None else None
            source_additional = Decimal(facts["additional_secured"]["parsed"]) if facts["additional_secured"]["parsed"] is not None else None
            if initial is None or current is None:
                reconciliation = {"value": None, "status": "missing_component", "difference": None}
            else:
                calculated = current - initial
                difference = calculated - source_additional if source_additional is not None else None
                status = "exact_match" if difference == 0 else ("tolerance_match" if difference is not None and abs(difference) <= Decimal("0.01") else "mismatch")
                reconciliation = {"value": str(calculated), "status": status, "difference": str(difference) if difference is not None else None}
            claims.append({
                "source_row": physical_row, "item_id": item_id, "raw_name": name,
                "display_name": name, "raw_status": text(row.get("Status")),
                "normalized_status": (text(row.get("Status")) or "unknown").lower(),
                "contractor": text(row.get("Contractor")), "carrier": text(row.get("Insurance Carrier")),
                "assignees": {"claim_handler": text(row.get("CH")), "admin": text(row.get("Admin")), "adjuster": text(row.get("Adjuster Name"))},
                "dates": {field: raw_date(row.get(field)) for field in ("Assigned", "Closed Date", "Date of Status Change", "Date of Loss")},
                "service_type": None, "property_type": None, "facts": facts,
                "reconciliation": reconciliation,
            })
        elif row_type == "subitem_header":
            subitem_headers.append({
                "source_row": physical_row, "parent_item_id": preceding_claim_id,
                "raw": row, "inference_method": "preceding_claim_position",
                "confidence": "limited",
            })
        else:
            subitem_details.append({
                "source_row": physical_row, "parent_item_id": preceding_claim_id,
                "header_source_row": preceding_header_row, "raw": row,
                "inference_method": "preceding_claim_and_header_position", "confidence": "limited",
            })

    claim_ids = {claim["item_id"] for claim in claims}
    updates, source_update_rows = [], []
    for physical_row, values in enumerate(
        updates_sheet.iter_rows(min_row=3, max_row=5959, max_col=11, values_only=True), 3
    ):
        row = row_object(values, update_headers)
        post_id = text(row.get("Post ID"))
        item_id = text(row.get("Item ID"))
        body = row.get("Update Content")
        timestamp = parse_update_timestamp(row.get("Created At"))
        record = {
            "source_row": physical_row, "post_id": post_id, "item_id": item_id,
            "matched": item_id in claim_ids, "body": text(body),
            "blank_body": text(body) is None, "body_fingerprint": fingerprint_body(body),
            "timestamp": timestamp, "author": text(row.get("User")),
            "raw": row,
        }
        updates.append(record)
        source_update_rows.append({
            "worksheet": UPDATES_SHEET, "physical_row": physical_row,
            "row_type": "update", "source_item_id": item_id, "raw": row, "normalized": row,
        })

    body_counts = Counter(update["body_fingerprint"] for update in updates if update["body_fingerprint"])
    for update in updates:
        update["duplicate_body"] = bool(update["body_fingerprint"] and body_counts[update["body_fingerprint"]] > 1)

    workbook_fingerprint = hashlib.sha256(
        f"{source_hash}:{IMPORTER_VERSION}:{len(source_rows)}:{len(updates)}".encode("utf-8")
    ).hexdigest()
    return {
        "source": {
            "system": "monday_archive", "type": "xlsx", "period": "2026-Q2",
            "filename": path.name, "sha256": source_hash, "importer_version": IMPORTER_VERSION,
            "transformation_version": TRANSFORMATION_VERSION, "fingerprint": workbook_fingerprint,
            "worksheets": [
                {"name": ARCHIVE_SHEET, "index": 0, "header_row": 5, "rows": 1726, "columns": 61, "headers": archive_headers},
                {"name": UPDATES_SHEET, "index": 1, "header_row": 2, "rows": 5959, "columns": 11, "headers": update_headers},
            ],
        },
        "source_rows": source_rows + source_update_rows,
        "claims": claims, "subitem_headers": subitem_headers,
        "subitem_details": subitem_details, "updates": updates,
    }


def summarize(payload: dict[str, Any]) -> dict[str, Any]:
    claims, updates = payload["claims"], payload["updates"]
    statuses = Counter(claim["raw_status"] for claim in claims)
    reconciliations = Counter(claim["reconciliation"]["status"] for claim in claims)
    unmatched = [update for update in updates if not update["matched"]]
    return {
        "source_filename": payload["source"]["filename"],
        "source_sha256": payload["source"]["sha256"],
        "run_fingerprint": payload["source"]["fingerprint"],
        "importer_version": IMPORTER_VERSION,
        "database_batch_plan": {
            "batch_count": 1 + sum(
                (len(payload[phase]) + size - 1) // size
                for phase, size in DATABASE_BATCH_SIZES.items()
            ),
            "batch_sizes": DATABASE_BATCH_SIZES,
            "finalizer": "count-gated",
        },
        "counts": {
            "claims": len(claims), "complete": statuses["Complete"], "closed": statuses["Closed"],
            "subitem_headers": len(payload["subitem_headers"]),
            "subitem_details": len(payload["subitem_details"]), "updates": len(updates),
            "unmatched_update_rows": len(unmatched),
            "unmatched_update_item_ids": len({u["item_id"] for u in unmatched}),
            "unique_post_ids": len({u["post_id"] for u in updates}),
            "blank_update_bodies": sum(u["blank_body"] for u in updates),
            "duplicate_body_surplus": sum(count - 1 for count in Counter(u["body_fingerprint"] for u in updates if u["body_fingerprint"]).values()),
            "exact_matches": reconciliations["exact_match"],
            "tolerance_matches": reconciliations["tolerance_match"],
            "mismatches": reconciliations["mismatch"],
            "missing_components": reconciliations["missing_component"],
        },
    }


def validate(summary: dict[str, Any]) -> list[str]:
    failures = []
    for key, expected in EXPECTED.items():
        actual = summary["counts"].get(key)
        if actual != expected:
            failures.append(f"{key}: expected {expected}, received {actual}")
    return failures


def database_batches(payload: dict[str, Any]):
    """Yield deterministic, bounded RPC payloads in dependency order."""
    sequence = 0
    empty = {key: [] for key in DATABASE_BATCH_SIZES}
    for phase, size in DATABASE_BATCH_SIZES.items():
        rows = payload[phase]
        for offset in range(0, len(rows), size):
            sequence += 1
            batch = {**empty, phase: rows[offset:offset + size]}
            yield {
                "source": payload["source"],
                **batch,
                "control": {
                    "phase": phase,
                    "sequence": sequence,
                    "offset": offset,
                    "batch_size": len(batch[phase]),
                    "finalize": False,
                },
            }
    sequence += 1
    yield {
        "source": payload["source"],
        **empty,
        "control": {
            "phase": "finalize",
            "sequence": sequence,
            "offset": 0,
            "batch_size": 0,
            "finalize": True,
        },
    }


def post_import_batch(url: str, key: str, batch: dict[str, Any]) -> dict[str, Any]:
    request = urllib.request.Request(
        f"{url.rstrip('/')}/rest/v1/rpc/import_q2_2026_archive",
        data=json.dumps({"payload": batch}, separators=(",", ":"), ensure_ascii=False).encode("utf-8"),
        headers={"apikey": key, "Authorization": f"Bearer {key}", "Content-Type": "application/json"},
        method="POST",
    )
    try:
        with urllib.request.urlopen(request, timeout=60) as response:
            return json.loads(response.read())
    except urllib.error.HTTPError as error:
        phase = batch["control"]["phase"]
        sequence = batch["control"]["sequence"]
        raise RuntimeError(
            f"Supabase import batch {sequence} ({phase}) failed ({error.code}): "
            f"{error.read().decode('utf-8')}"
        ) from error


def mark_import_failed(url: str, key: str, payload: dict[str, Any], error: Exception) -> None:
    request = urllib.request.Request(
        f"{url.rstrip('/')}/rest/v1/rpc/mark_q2_2026_import_failed",
        data=json.dumps({
            "workbook_sha256": payload["source"]["sha256"],
            "version": payload["source"]["importer_version"],
            "failure_message": str(error)[:2000],
        }, separators=(",", ":")).encode("utf-8"),
        headers={"apikey": key, "Authorization": f"Bearer {key}", "Content-Type": "application/json"},
        method="POST",
    )
    try:
        with urllib.request.urlopen(request, timeout=15):
            return
    except Exception:
        # A running job still unambiguously signals interruption if the marker cannot be reached.
        return


def confirm_import_target(confirm_target: str | None) -> tuple[str, str, str]:
    configured = os.environ.get("TOTALSCOPE_IMPORT_TARGET")
    url = os.environ.get("SUPABASE_URL", "")
    if configured not in {"local", "staging"}:
        raise RuntimeError("TOTALSCOPE_IMPORT_TARGET must explicitly be local or staging; production targets are prohibited")
    if confirm_target != configured:
        raise RuntimeError(f"Import confirmation mismatch: expected --confirm-target {configured}")
    parsed = urllib.parse.urlparse(url)
    hostname = (parsed.hostname or "").lower()
    is_local = hostname in {"127.0.0.1", "localhost", "::1"}
    if configured == "local" and not is_local:
        raise RuntimeError("TOTALSCOPE_IMPORT_TARGET=local requires a loopback Supabase URL")
    if configured == "staging" and ("prod" in hostname or "production" in hostname):
        raise RuntimeError("Production-like target hostname rejected")
    return configured, hostname or "unresolved", "local" if is_local else "hosted"


def import_payload(payload: dict[str, Any], confirm_target: str | None) -> dict[str, Any]:
    url = os.environ.get("SUPABASE_URL")
    key = os.environ.get("SUPABASE_SERVICE_ROLE_KEY")
    if not url or not key:
        raise RuntimeError("SUPABASE_URL and SUPABASE_SERVICE_ROLE_KEY are required for database import")
    target, hostname, location = confirm_import_target(confirm_target)
    print(f"Confirmed import target: environment={target} location={location} host={hostname}", file=sys.stderr)
    batches = list(database_batches(payload))
    final_result = {}
    try:
        for index, batch in enumerate(batches, 1):
            control = batch["control"]
            print(
                f"Database batch {index}/{len(batches)}: {control['phase']} "
                f"offset={control['offset']} rows={control['batch_size']}",
                file=sys.stderr,
            )
            final_result = post_import_batch(url, key, batch)
            if final_result.get("status") == "failed":
                raise RuntimeError(f"Supabase import batch failed: {final_result}")
    except Exception as error:
        mark_import_failed(url, key, payload, error)
        raise
    return {
        **final_result,
        "batch_count": len(batches),
        "batch_sizes": DATABASE_BATCH_SIZES,
        "resume_behavior": "conflict-safe retry by source/importer fingerprint",
    }


def validate_database() -> dict[str, Any]:
    url = os.environ.get("SUPABASE_URL")
    key = os.environ.get("SUPABASE_SERVICE_ROLE_KEY")
    if not url or not key:
        return {"status": "skipped", "reason": "Supabase environment variables are not configured"}
    request = urllib.request.Request(
        f"{url.rstrip('/')}/rest/v1/q2_2026_import_validation?select=*&order=import_job_id.desc&limit=1",
        headers={"apikey": key, "Authorization": f"Bearer {key}"},
    )
    try:
        with urllib.request.urlopen(request, timeout=30) as response:
            rows = json.loads(response.read())
    except urllib.error.HTTPError as error:
        raise RuntimeError(f"Database validation failed ({error.code}): {error.read().decode('utf-8')}") from error
    if not rows:
        raise RuntimeError("Database validation found no Q2 2026 import")
    row = rows[0]
    mapping = {
        "claim_count": "claims", "complete_status_count": "complete", "closed_status_count": "closed",
        "staged_subitem_header_count": "subitem_headers", "staged_subitem_detail_count": "subitem_details",
        "update_count": "updates", "unmatched_update_row_count": "unmatched_update_rows",
        "unmatched_update_item_id_count": "unmatched_update_item_ids", "unique_post_id_count": "unique_post_ids",
        "additional_rcv_exact_match_count": "exact_matches",
        "additional_rcv_tolerance_only_count": "tolerance_matches",
        "additional_rcv_mismatch_count": "mismatches",
        "additional_rcv_missing_component_count": "missing_components",
    }
    failures = [
        f"{column}: expected {EXPECTED[key]}, received {row.get(column)}"
        for column, key in mapping.items() if int(row.get(column, -1)) != EXPECTED[key]
    ]
    return {"status": "pass" if not failures else "fail", "failures": failures, "actual": row}


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("command", choices=("inspect", "import", "validate"))
    parser.add_argument("--allow-checksum-override", action="store_true")
    parser.add_argument("--json-output", type=Path)
    parser.add_argument("--confirm-target", choices=("local", "staging"))
    args = parser.parse_args()
    try:
        before = hashlib.sha256(SOURCE.read_bytes()).hexdigest() if SOURCE.exists() else None
        payload = parse_workbook(allow_checksum_override=args.allow_checksum_override)
        summary = summarize(payload)
        failures = validate(summary)
        if args.command == "import":
            summary["database"] = import_payload(payload, args.confirm_target)
        elif args.command == "validate":
            summary["database_validation"] = validate_database()
            failures.extend(summary["database_validation"].get("failures", []))
        after = hashlib.sha256(SOURCE.read_bytes()).hexdigest()
        summary["workbook_unchanged"] = before == after == payload["source"]["sha256"]
        summary["validation"] = {"status": "pass" if not failures else "fail", "failures": failures}
        encoded = json.dumps(summary, indent=2, sort_keys=True)
        print(encoded)
        if args.json_output:
            args.json_output.parent.mkdir(parents=True, exist_ok=True)
            args.json_output.write_text(encoded + "\n", encoding="utf-8")
        return 0 if not failures else 1
    except Exception as error:
        print(json.dumps({"validation": {"status": "fail", "failures": [str(error)]}}, indent=2), file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
